import psycopg2
import stdiomask
from openpyxl import load_workbook
from psycopg2 import OperationalError  # NOQA (PyCharm ругается, при это класс рабочий)


def creating_list_of_points(sheet):
    """ Функция принимает на вход лист Excel документа, в котором гарантированно есть два столбца.
    На выходе функции получаем список, элементы которого картежи из значений столбцов endpoint_id и endpoint_name.
    """
    i = 2
    list_point = list()
    while sheet.cell(row=i, column=1).value is not None:
        list_point.append(
            (sheet.cell(row=i, column=1).value, sheet.cell(row=i, column=2).value)
        )
        i += 1
    return list_point


def create_connection(db_password, db_name="postgres", db_user="postgres", db_host="127.0.0.1", db_port="5432"):
    """ Функция подключения к БД."""
    connection = None
    try:
        connection = psycopg2.connect(
            database=db_name,
            user=db_user,
            password=db_password,
            host=db_host,
            port=db_port,
        )
        print("Connection to PostgreSQL DB successful")
    except OperationalError as e:
        print(f"The error '{e}' occurred")
    return connection


def execute_query(connection, query):
    """ Функция отправки запроса в БД. Выводит на экран результат транзакции (успех или ошибка). """
    connection.autocommit = True
    cursor = connection.cursor()
    try:
        cursor.execute(query)
        print("Query executed successfully")
    except OperationalError as e:
        print(f"The error '{e}' occurred")


def loading_or_updating_data_in_database():
    """ Функция объединяющая в себе создание таблицы и добавление/обновление записей в существующую таблицу"""
    create_endpoint_names_table = """
    CREATE TABLE IF NOT EXISTS endpoint_names (
        endpoint_id INTEGER PRIMARY KEY,
        endpoint_name TEXT
    );
    """
    execute_query(our_connection, create_endpoint_names_table)

    endpoint_names_records = ", ".join(["%s"] * len(new_list))

    insert_query = (
        f"INSERT INTO endpoint_names (endpoint_id, endpoint_name) VALUES {endpoint_names_records} \
        ON CONFLICT(endpoint_id) DO UPDATE SET endpoint_id=EXCLUDED.endpoint_id, endpoint_name=EXCLUDED.endpoint_name;"
    )

    our_connection.autocommit = True
    cursor = our_connection.cursor()
    cursor.execute(insert_query, new_list)
    pass


way_to_file = input('Введите путь до файла "названия точек.xlsm": ')
wb = load_workbook(way_to_file)
our_sheet = wb[wb.sheetnames[0]]

new_list = creating_list_of_points(our_sheet)

password = stdiomask.getpass(prompt='Введите пароль от БД: ')
our_connection = create_connection(password, "postgres", "postgres", "127.0.0.1", "5432")
loading_or_updating_data_in_database()
